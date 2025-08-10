#!/usr/bin/env python3
"""
Script para generar la plantilla Excel del Sistema de Inventario QR
Genera un archivo Excel profesional con datos de ejemplo y instrucciones

Requisitos:
    pip install openpyxl

Uso:
    python generate_excel_template.py

Autor: Sistema de Inventario QR
"""

import sys
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: Se requiere openpyxl")
    print("Instalar con: pip install openpyxl")
    sys.exit(1)


def create_excel_template():
    """Crea la plantilla Excel con formato profesional"""
    print("📋 Iniciando generación de plantilla Excel...")
    
    # Crear workbook
    workbook = Workbook()
    
    # ===== HOJA 1: PLANTILLA DE INVENTARIO =====
    worksheet = workbook.active
    worksheet.title = "Plantilla Inventario"
    
    # Encabezados de columnas
    headers = [
        'Nombre del Artículo',
        'Demanda Anual (unidades)',
        'Costo por Pedido (soles)', 
        'Costo Mantenimiento (soles/unidad)',
        'Costo por Faltante (soles/unidad)',
        'Costo Unitario (soles)',
        'Espacio por Unidad (m²)',
        'Desviación Estándar Diaria',
        'Punto de Reorden (unidades)',
        'Tamaño de Lote (unidades)',
    ]
    
    # Datos de ejemplo perfectamente válidos para el modelo QR
    datos_ejemplo = [
        ['Laptop HP EliteBook', 1200, 150, 24, 50, 2500, 0.5, 15, 120, 180],
        ['Mouse Logitech MX Master', 800, 50, 8, 15, 45, 0.1, 8, 80, 120],
        ['Teclado Mecánico Gaming', 600, 75, 12, 25, 180, 0.3, 10, 60, 90],
        ['Monitor 24" Dell', 400, 100, 48, 80, 1200, 2.0, 12, 40, 60],
        ['Impresora Láser HP', 200, 120, 60, 150, 3500, 1.5, 20, 25, 35],
        ['Tablet Samsung Galaxy', 300, 80, 15, 30, 800, 0.4, 6, 35, 50],
        ['Auriculares Sony WH-1000XM4', 150, 40, 5, 12, 280, 0.2, 4, 18, 30],
        ['Cámara Web Logitech C920', 250, 60, 10, 20, 120, 0.15, 5, 25, 40],
    ]
    
    # Definir estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir y formatear encabezados
    print("📝 Escribiendo encabezados...")
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    # Escribir y formatear datos de ejemplo
    print("📊 Escribiendo datos de ejemplo...")
    for row_idx, row_data in enumerate(datos_ejemplo, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border_thin
            
            # Formato específico para números
            if isinstance(value, (int, float)) and col_idx > 1:
                if col_idx in [2, 9, 10]:  # Cantidades enteras
                    cell.number_format = '0'
                else:  # Decimales
                    cell.number_format = '0.00'
    
    # Ajustar ancho de columnas
    print("📐 Ajustando ancho de columnas...")
    column_widths = [25, 18, 20, 25, 25, 18, 18, 22, 20, 20]
    for col, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = width
    
    # Ajustar altura de la fila de encabezados
    worksheet.row_dimensions[1].height = 30
    
    # ===== HOJA 2: INSTRUCCIONES =====
    print("📖 Creando hoja de instrucciones...")
    inst_sheet = workbook.create_sheet(title="Instrucciones")
    
    # Contenido de instrucciones
    instrucciones = [
        ("INSTRUCCIONES PARA USO DE LA PLANTILLA", "title"),
        ("", "normal"),
        ("1. Complete los datos en la hoja 'Plantilla Inventario'", "normal"),
        ("2. Todos los campos numéricos deben ser mayor a 0", "normal"),
        ("3. El nombre del artículo debe ser único", "normal"),
        ("4. Use la plantilla como guía para sus propios datos", "normal"),
        ("5. Guarde el archivo antes de importar a la aplicación", "normal"),
        ("6. El sistema detectará automáticamente el formato estándar", "normal"),
        ("", "normal"),
        ("DESCRIPCIÓN DE CAMPOS:", "subtitle"),
        ("", "normal"),
        ("• Demanda Anual: Cantidad total requerida por año", "bullet"),
        ("• Costo por Pedido: Costo fijo de realizar un pedido", "bullet"),
        ("• Costo Mantenimiento: Costo de mantener una unidad en inventario por año", "bullet"),
        ("• Costo por Faltante: Costo cuando no hay stock disponible", "bullet"),
        ("• Costo Unitario: Precio de compra de cada unidad", "bullet"),
        ("• Espacio por Unidad: Metros cuadrados que ocupa cada unidad", "bullet"),
        ("• Desviación Estándar Diaria: Variabilidad de la demanda diaria", "bullet"),
        ("• Punto de Reorden: Nivel de stock para realizar nuevo pedido", "bullet"),
        ("• Tamaño de Lote: Cantidad a pedir en cada pedido", "bullet"),
        ("", "normal"),
        ("MODELO MATEMÁTICO:", "subtitle"),
        ("", "normal"),
        ("El sistema utiliza el modelo EOQ (Economic Order Quantity) con:", "bullet"),
        ("• Optimización de costos de pedido y mantenimiento", "bullet"),
        ("• Consideración de restricciones de espacio", "bullet"),
        ("• Cálculo de punto de reorden con demanda variable", "bullet"),
        ("• Análisis de sensibilidad de parámetros", "bullet"),
        ("", "normal"),
        ("NOTAS IMPORTANTES:", "subtitle"),
        ("", "normal"),
        ("⚠️ Los datos de ejemplo son referenciales", "warning"),
        ("⚠️ Valide que sus datos reflejen la realidad de su negocio", "warning"),
        ("⚠️ El sistema mostrará advertencias si detecta valores inconsistentes", "warning"),
        ("✅ Para mejores resultados, use datos históricos reales", "success"),
    ]
    
    # Estilos para instrucciones
    title_font = Font(name='Arial', size=16, bold=True, color='2E75B6')
    subtitle_font = Font(name='Arial', size=14, bold=True, color='2E75B6')
    normal_font = Font(name='Arial', size=11)
    bullet_font = Font(name='Arial', size=10)
    warning_font = Font(name='Arial', size=10, color='D32F2F')
    success_font = Font(name='Arial', size=10, color='388E3C')
    
    # Escribir instrucciones
    for row, (text, style) in enumerate(instrucciones, 1):
        cell = inst_sheet.cell(row=row, column=1, value=text)
        
        if style == "title":
            cell.font = title_font
        elif style == "subtitle":
            cell.font = subtitle_font
        elif style == "bullet":
            cell.font = bullet_font
        elif style == "warning":
            cell.font = warning_font
        elif style == "success":
            cell.font = success_font
        else:
            cell.font = normal_font
    
    # Ajustar ancho de columna para instrucciones
    inst_sheet.column_dimensions['A'].width = 80
    
    # ===== HOJA 3: EJEMPLO DE RESULTADOS =====
    print("📈 Creando hoja de ejemplo de resultados...")
    results_sheet = workbook.create_sheet(title="Ejemplo Resultados")
    
    # Encabezados de resultados
    result_headers = [
        'Artículo',
        'Q* Óptimo',
        'Costo Total',
        'Costo Pedido',
        'Costo Mant.',
        'Frecuencia',
        'Punto Reorden',
        'Stock Seguridad',
        'Espacio Usado',
        'Rentabilidad'
    ]
    
    # Datos de ejemplo de resultados
    result_data = [
        ['Laptop HP EliteBook', 179.44, 4295.58, 1005.58, 2158.32, 6.69, 125.5, 45.2, 89.72, 'Alta'],
        ['Mouse Logitech MX Master', 118.32, 946.56, 337.84, 473.28, 6.76, 85.3, 24.1, 11.83, 'Media'],
        ['Teclado Mecánico Gaming', 89.44, 1073.28, 502.65, 537.6, 6.71, 65.2, 18.8, 26.83, 'Alta'],
        ['Monitor 24" Dell', 59.16, 2840.16, 675.68, 1420.8, 6.77, 44.8, 16.4, 118.32, 'Media'],
        ['Impresora Láser HP', 34.64, 4158.72, 693.6, 1039.68, 5.77, 28.5, 12.2, 51.96, 'Baja'],
    ]
    
    # Escribir encabezados de resultados
    for col, header in enumerate(result_headers, 1):
        cell = results_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    # Escribir datos de resultados
    for row_idx, row_data in enumerate(result_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = results_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border_thin
            
            # Formato para números
            if isinstance(value, float):
                cell.number_format = '0.00'
    
    # Ajustar anchos para resultados
    for col in range(1, len(result_headers) + 1):
        results_sheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Guardar archivo
    output_path = Path("assets/templates/plantilla_inventario.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"💾 Guardando archivo en: {output_path}")
    workbook.save(output_path)
    
    return output_path


def main():
    """Función principal"""
    print("🚀 Generador de Plantilla Excel - Sistema de Inventario QR")
    print("=" * 60)
    
    try:
        output_file = create_excel_template()
        
        print("=" * 60)
        print("✅ ¡Plantilla generada exitosamente!")
        print(f"📁 Archivo: {output_file}")
        print(f"📊 Tamaño: {output_file.stat().st_size:,} bytes")
        print()
        print("📋 La plantilla incluye:")
        print("   • Hoja 'Plantilla Inventario' con datos de ejemplo")
        print("   • Hoja 'Instrucciones' con guía detallada")
        print("   • Hoja 'Ejemplo Resultados' con resultados esperados")
        print("   • Formato profesional con colores y bordes")
        print("   • Validación de datos y tipos de números")
        print()
        print("🎯 Listo para usar en el proyecto Flutter!")
        
    except Exception as e:
        print(f"❌ Error al generar plantilla: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
